"""刷单清单：模板下载、导入校验、写入 config.review_orders。"""

from __future__ import annotations

import io
import re

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models import DataSource
from app.services.ds_settings import get_ds_config, save_ds_config
from app.services.fact_provider import load_fact_rows
from app.services.field_aggregator import ORDER_ID_CANDIDATES, SKU_ID_CANDIDATES, _extract, _to_number

# (内部键, 模板列名, 是否必填) — 物流费请用「运费模板」单独导入
REVIEW_COLUMNS: list[tuple[str, str, bool]] = [
    ("order_id", "Order ID", True),
    ("sku_id", "SKU ID", True),
    ("amount", "刷单金额", False),
    ("commission", "刷单佣金", False),
    ("service_fee", "刷单服务费", False),
    ("cost", "刷单成本", False),
]

# 刷单运费单独导入 → 日报「刷单物流费用」
REVIEW_LOGISTICS_IMPORT_COLUMNS: list[tuple[str, str, bool]] = [
    ("order_id", "Order ID", True),
    ("sku_id", "SKU ID", True),
    ("logistics", "刷单运费", False),
]

REVIEW_LOGISTICS_TEMPLATE_HEADERS = [col[1] for col in REVIEW_LOGISTICS_IMPORT_COLUMNS]

TEMPLATE_HEADERS = [col[1] for col in REVIEW_COLUMNS]

REVIEW_FIELD_CODES = {
    "amount": "mc_review_amount",
    "commission": "mc_review_commission",
    "service_fee": "mc_review_service_fee",
    "logistics": "mc_review_logistics",
    "cost": "mc_review_cost",
}

_HEADER_ALIASES: dict[str, str] = {}
for key, header, _ in REVIEW_COLUMNS:
    _HEADER_ALIASES[header.lower()] = key
    _HEADER_ALIASES[header.replace(" ", "").lower()] = key
for key, header, _ in REVIEW_LOGISTICS_IMPORT_COLUMNS:
    _HEADER_ALIASES[header.lower()] = key
    _HEADER_ALIASES[header.replace(" ", "").lower()] = key
    if key == "logistics":
        _HEADER_ALIASES["运费"] = key
        _HEADER_ALIASES["物流费"] = key
        _HEADER_ALIASES["刷单运费"] = key
        _HEADER_ALIASES["刷单物流费"] = key
        _HEADER_ALIASES["刷单物流费用"] = key
_HEADER_ALIASES["order id"] = "order_id"
_HEADER_ALIASES["skuid"] = "sku_id"


def build_review_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "刷单清单"
    ws.append([
        "说明：每行一条刷单 SKU；Order ID、SKU ID 必填；金额/佣金/服务费/成本填在对应列；刷单运费请用「运费模板」单独导入"
    ])
    ws.append(TEMPLATE_HEADERS)
    ws.append(["1234567890123456789", "9876543210", 100, 10, 5, 50])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_review_logistics_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "刷单运费"
    ws.append([
        "说明：每行一条刷单 SKU 的运费；Order ID、SKU ID 必填；刷单运费填在对应列；导入后汇总至日报「刷单物流费用」，同 Order+SKU 覆盖"
    ])
    ws.append(REVIEW_LOGISTICS_TEMPLATE_HEADERS)
    ws.append(["1234567890123456789", "9876543210", 1.5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_header(cell) -> str:
    return re.sub(r"\s+", " ", str(cell or "").strip())


def _header_to_key(header: str) -> str | None:
    h = _normalize_header(header).lower()
    return _HEADER_ALIASES.get(h) or _HEADER_ALIASES.get(h.replace(" ", ""))


def _known_order_sku_keys(db: Session, ds: DataSource) -> set[tuple[str, str]]:
    cfg = get_ds_config(ds)
    store = (cfg.get("meta") or {}).get("店铺名称") or ds.name
    rows, _ = load_fact_rows(db, ds.id, store)
    order_sheet = cfg.get("order_sheet")
    order_id_col = cfg.get("order_id_col") or "Order ID"
    sku_id_col = cfg.get("sku_id_col") or "SKU ID"
    keys: set[tuple[str, str]] = set()
    for r in rows:
        if order_sheet and r.sheet_name != order_sheet:
            continue
        oid = _extract(r.row_data, [order_id_col, *ORDER_ID_CANDIDATES])
        sku = _extract(r.row_data, [sku_id_col, *SKU_ID_CANDIDATES])
        if oid and sku:
            keys.add((oid, sku))
    return keys


def parse_review_upload(content: bytes) -> tuple[list[dict], list[str]]:
    """返回 (review_rows, errors)。"""
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return [], ["无法读取 Excel 文件，请使用 .xlsx 格式"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["文件为空"]

    header_row_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows[:15]):
        cells = [_normalize_header(c) for c in row]
        mapping: dict[str, int] = {}
        for j, h in enumerate(cells):
            key = _header_to_key(h)
            if key and key not in mapping:
                mapping[key] = j
        if "order_id" in mapping and "sku_id" in mapping:
            header_row_idx = i
            col_map = mapping
            break

    if header_row_idx is None:
        return [], ["缺少必填列「Order ID」和「SKU ID」（或对应中文列头）"]

    parsed: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for line_no, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(key: str):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        oid = str(cell("order_id") or "").strip()
        sku = str(cell("sku_id") or "").strip()
        if not oid and not sku:
            continue
        if oid.startswith("1234567890") and sku == "9876543210":
            continue  # 跳过模板示例行

        if not oid:
            errors.append(f"第 {line_no} 行：Order ID 不能为空")
            continue
        if not sku:
            errors.append(f"第 {line_no} 行：SKU ID 不能为空")
            continue

        pair = (oid, sku)
        if pair in seen:
            errors.append(f"第 {line_no} 行：Order ID + SKU ID 重复（{oid} / {sku}）")
            continue
        seen.add(pair)

        record: dict = {"order_id": oid, "sku_id": sku}
        label_by_key = {k: h for k, h, _ in REVIEW_COLUMNS}
        for key, _, _required in REVIEW_COLUMNS:
            if key in ("order_id", "sku_id"):
                continue
            raw = cell(key)
            if raw is None or str(raw).strip() == "":
                record[key] = 0.0
                continue
            num = _to_number(raw)
            text = str(raw).strip()
            if text and text not in ("0", "0.0") and num == 0.0 and not isinstance(raw, (int, float)):
                errors.append(f"第 {line_no} 行：{label_by_key[key]} 不是有效数字")
                continue
            record[key] = num
        parsed.append(record)

    if not parsed and not errors:
        errors.append("未解析到有效刷单数据")
    return parsed, errors


def parse_review_logistics_upload(content: bytes) -> tuple[list[dict], list[str]]:
    """解析运费导入 Excel，返回 (logistics_rows, errors)。"""
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return [], ["无法读取 Excel 文件，请使用 .xlsx 格式"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["文件为空"]

    header_row_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows[:15]):
        cells = [_normalize_header(c) for c in row]
        mapping: dict[str, int] = {}
        for j, h in enumerate(cells):
            key = _header_to_key(h)
            if key and key not in mapping:
                mapping[key] = j
        if "order_id" in mapping and "sku_id" in mapping and "logistics" in mapping:
            header_row_idx = i
            col_map = mapping
            break

    if header_row_idx is None:
        return [], ["缺少必填列「Order ID」「SKU ID」与「刷单运费」（或对应中文列头）"]

    parsed: list[dict] = []
    seen: set[tuple[str, str]] = set()
    label_by_key = {k: h for k, h, _ in REVIEW_LOGISTICS_IMPORT_COLUMNS}
    for line_no, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(key: str):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        oid = str(cell("order_id") or "").strip()
        sku = str(cell("sku_id") or "").strip()
        if not oid and not sku:
            continue
        if oid.startswith("1234567890") and sku == "9876543210":
            continue

        if not oid:
            errors.append(f"第 {line_no} 行：Order ID 不能为空")
            continue
        if not sku:
            errors.append(f"第 {line_no} 行：SKU ID 不能为空")
            continue

        pair = (oid, sku)
        if pair in seen:
            errors.append(f"第 {line_no} 行：Order ID + SKU ID 重复（{oid} / {sku}）")
            continue
        seen.add(pair)

        raw = cell("logistics")
        if raw is None or str(raw).strip() == "":
            logistics = 0.0
        else:
            num = _to_number(raw)
            text = str(raw).strip()
            if text and text not in ("0", "0.0") and num == 0.0 and not isinstance(raw, (int, float)):
                errors.append(f"第 {line_no} 行：{label_by_key['logistics']} 不是有效数字")
                continue
            logistics = num
        parsed.append({"order_id": oid, "sku_id": sku, "logistics": logistics})

    if not parsed and not errors:
        errors.append("未解析到有效运费数据")
    return parsed, errors


def _merge_review_logistics(existing: list[dict], logistics_rows: list[dict]) -> list[dict]:
    merged: dict[tuple[str, str], dict] = {}
    for row in existing:
        oid = str(row.get("order_id", "")).strip()
        sku = str(row.get("sku_id", "")).strip()
        if oid and sku:
            merged[(oid, sku)] = dict(row)
    for row in logistics_rows:
        oid = row["order_id"]
        sku = row["sku_id"]
        base = merged.get((oid, sku)) or {
            "order_id": oid,
            "sku_id": sku,
            "amount": 0.0,
            "commission": 0.0,
            "service_fee": 0.0,
            "cost": 0.0,
        }
        base["logistics"] = row.get("logistics", 0.0)
        merged[(oid, sku)] = base
    return list(merged.values())


REVIEW_LOGISTICS_MODE_FIXED = "per_order_fixed"
REVIEW_LOGISTICS_MODE_IMPORT = "from_import"
DEFAULT_REVIEW_LOGISTICS_PER_ORDER = 1.0


def review_logistics_mode(cfg: dict) -> str:
    mode = (cfg.get("review_logistics_mode") or "").strip()
    if mode == REVIEW_LOGISTICS_MODE_IMPORT:
        return REVIEW_LOGISTICS_MODE_IMPORT
    return REVIEW_LOGISTICS_MODE_FIXED


def review_logistics_per_order(cfg: dict) -> float:
    raw = cfg.get("review_logistics_per_order")
    if raw is None or str(raw).strip() == "":
        return DEFAULT_REVIEW_LOGISTICS_PER_ORDER
    return max(0.0, _to_number(raw))


def review_logistics_exclude_same_day_refund(cfg: dict) -> bool:
    return bool(cfg.get("review_logistics_exclude_same_day_refund"))


def distinct_review_order_count(
    reviews: list[dict],
    exclude_order_ids: set[str] | None = None,
) -> int:
    exclude = exclude_order_ids or set()
    return len({
        str(r.get("order_id", "")).strip()
        for r in reviews
        if str(r.get("order_id", "")).strip() and str(r.get("order_id", "")).strip() not in exclude
    })


def review_logistics_total(
    reviews: list[dict],
    cfg: dict,
    *,
    exclude_same_day_order_ids: set[str] | None = None,
) -> float:
    if review_logistics_mode(cfg) == REVIEW_LOGISTICS_MODE_IMPORT:
        exclude: set[str] = set()
        if review_logistics_exclude_same_day_refund(cfg) and exclude_same_day_order_ids:
            exclude = exclude_same_day_order_ids
        total = 0.0
        for row in reviews:
            oid = str(row.get("order_id", "")).strip()
            if oid and oid in exclude:
                continue
            total += _to_number(row.get("logistics"))
        return total
    exclude: set[str] = set()
    if review_logistics_exclude_same_day_refund(cfg) and exclude_same_day_order_ids:
        exclude = exclude_same_day_order_ids
    order_count = distinct_review_order_count(reviews, exclude)
    return order_count * review_logistics_per_order(cfg)


def review_logistics_rule_summary(cfg: dict) -> str:
    reviews = cfg.get("review_orders") or []
    if review_logistics_mode(cfg) == REVIEW_LOGISTICS_MODE_IMPORT:
        total = review_logistics_total(reviews, cfg)
        suffix = " · 排除当日退单" if review_logistics_exclude_same_day_refund(cfg) else ""
        return f"Excel 导入运费合计 ${total:g}{suffix} · {len(reviews)} 行"
    order_count = distinct_review_order_count(reviews)
    per_order = review_logistics_per_order(cfg)
    suffix = " · 排除当日退单" if review_logistics_exclude_same_day_refund(cfg) else ""
    return f"按单固定 ${per_order:g}/单 × {order_count} 单刷单订单{suffix}"


def review_import_stats(
    reviews: list[dict],
    cfg: dict,
    *,
    exclude_same_day_order_ids: set[str] | None = None,
) -> dict:
    exclude: set[str] = set()
    if review_logistics_exclude_same_day_refund(cfg) and exclude_same_day_order_ids:
        exclude = exclude_same_day_order_ids
    order_count = distinct_review_order_count(reviews, exclude)
    row_count = len(reviews)
    mode = review_logistics_mode(cfg)
    logistics_total = review_logistics_total(
        reviews,
        cfg,
        exclude_same_day_order_ids=exclude_same_day_order_ids,
    )
    return {
        "row_count": row_count,
        "order_count": order_count,
        "logistics_total": logistics_total,
        "logistics_mode": mode,
        "logistics_per_order": review_logistics_per_order(cfg),
    }


def review_field_values(
    ds_config: dict,
    same_day_refund_order_ids: set[str] | None = None,
) -> dict[str, float]:
    reviews = ds_config.get("review_orders") or []
    out = {code: 0.0 for code in REVIEW_FIELD_CODES.values()}
    for row in reviews:
        for key, code in REVIEW_FIELD_CODES.items():
            if key == "logistics":
                continue
            out[code] += _to_number(row.get(key))
    stats = review_import_stats(
        reviews,
        ds_config,
        exclude_same_day_order_ids=same_day_refund_order_ids,
    )
    out[REVIEW_FIELD_CODES["logistics"]] = stats["logistics_total"]
    return out


def review_order_ids_from_rows(rows: list[dict]) -> list[str]:
    return sorted({str(r["order_id"]).strip() for r in rows if r.get("order_id")})


def import_review_orders(
    db: Session,
    ds: DataSource,
    content: bytes,
    *,
    strict: bool = True,
) -> dict:
    review_rows, parse_errors = parse_review_upload(content)
    if parse_errors and not review_rows:
        return {"ok": False, "errors": parse_errors, "imported": 0}

    known = _known_order_sku_keys(db, ds)
    unknown = [
        f"{r['order_id']}/{r['sku_id']}"
        for r in review_rows
        if (r["order_id"], r["sku_id"]) not in known
    ]
    errors = list(parse_errors)
    if strict and unknown:
        preview = "、".join(unknown[:5])
        suffix = f" 等 {len(unknown)} 条" if len(unknown) > 5 else ""
        errors.append(f"以下 Order ID + SKU ID 在订单主表中不存在：{preview}{suffix}")

    if errors:
        return {"ok": False, "errors": errors, "imported": 0, "unknown_count": len(unknown)}

    order_ids = review_order_ids_from_rows(review_rows)
    cfg = save_ds_config(db, ds, {
        "review_orders": review_rows,
        "review_order_ids": order_ids,
    })
    stats = review_import_stats(review_rows, cfg)
    return {
        "ok": True,
        "imported": len(review_rows),
        "review_order_count": len(cfg.get("review_orders") or []),
        "review_order_distinct": stats["order_count"],
        "review_logistics_total": stats["logistics_total"],
        "review_logistics_summary": review_logistics_rule_summary(cfg),
    }


def import_review_logistics(
    db: Session,
    ds: DataSource,
    content: bytes,
    *,
    strict: bool = True,
) -> dict:
    logistics_rows, parse_errors = parse_review_logistics_upload(content)
    if parse_errors and not logistics_rows:
        return {"ok": False, "errors": parse_errors, "imported": 0}

    known = _known_order_sku_keys(db, ds)
    unknown = [
        f"{r['order_id']}/{r['sku_id']}"
        for r in logistics_rows
        if (r["order_id"], r["sku_id"]) not in known
    ]
    errors = list(parse_errors)
    if strict and unknown:
        preview = "、".join(unknown[:5])
        suffix = f" 等 {len(unknown)} 条" if len(unknown) > 5 else ""
        errors.append(f"以下 Order ID + SKU ID 在订单主表中不存在：{preview}{suffix}")

    if errors:
        return {"ok": False, "errors": errors, "imported": 0, "unknown_count": len(unknown)}

    cfg = get_ds_config(ds)
    merged = _merge_review_logistics(cfg.get("review_orders") or [], logistics_rows)
    order_ids = review_order_ids_from_rows(merged)
    cfg = save_ds_config(db, ds, {
        "review_orders": merged,
        "review_order_ids": order_ids,
        "review_logistics_mode": REVIEW_LOGISTICS_MODE_IMPORT,
    })
    stats = review_import_stats(merged, cfg)
    return {
        "ok": True,
        "imported": len(logistics_rows),
        "review_order_count": len(cfg.get("review_orders") or []),
        "review_order_distinct": stats["order_count"],
        "review_logistics_total": stats["logistics_total"],
        "review_logistics_summary": review_logistics_rule_summary(cfg),
    }
