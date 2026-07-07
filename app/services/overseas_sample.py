"""生成模拟海外电商 RPA 导出 Excel（多 Sheet、英文列头、真实业务维度）"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OVERSEAS_STORE_META = {
    "platform": "TikTok Shop",
    "region": "UK/EU",
    "store_code": "TT-UK-017",
    "currency": "USD",
    "report_date": "2025-06-23",
}

# (Date, Order ID, SKU, Title, Units, Gross, SKU Discount, Order Discount, Commission, Shipping, Refund, Country, Status, Fulfillment)
ORDER_ROWS = [
    ("2025-06-23", "5791023847291", "B0TK8X1A", "Wireless Earbuds Pro ANC", 1, 49.99, 5.00, 8.00, 7.50, 0, 0, "UK", "Delivered", "FBT"),
    ("2025-06-23", "5791023847291", "B0TK8X2B", "USB-C Cable 2m", 2, 25.98, 2.60, 8.00, 3.90, 0, 0, "UK", "Delivered", "FBT"),
    ("2025-06-23", "5791023847291", "B0TK8X3C", "Earbuds Case", 1, 12.99, 0, 8.00, 1.95, 0, 0, "UK", "Delivered", "FBT"),
    ("2025-06-23", "5791023847292", "B0TK8X4D", "Portable Blender", 1, 35.00, 3.50, 0, 5.25, 3.50, 0, "UK", "Delivered", "Merchant"),
    ("2025-06-23", "5791023847293", "B0TK8X5E", "LED Desk Lamp", 2, 55.98, 0, 6.00, 8.40, 0, 0, "DE", "Delivered", "FBT"),
    ("2025-06-23", "5791023847293", "B0TK8X6F", "Lamp Bulb Pack", 1, 9.99, 1.00, 6.00, 1.50, 0, 0, "DE", "Delivered", "FBT"),
    ("2025-06-23", "5791023847294", "B0TK8X7G", "Water Bottle 750ml", 4, 63.96, 6.40, 0, 9.60, 0, 0, "UK", "Shipped", "FBT"),
    ("2025-06-23", "5791023847300", "B0TK8X8H", "Electric Toothbrush", 1, 42.99, 4.30, 5.00, 6.45, 0, 0, "UK", "Delivered", "FBT"),
    ("2025-06-23", "5791023847301", "B0TK8X9J", "Car Phone Mount", 2, 29.98, 0, 0, 4.50, 0, 0, "UK", "Delivered", "FBT"),
    ("2025-06-23", "5791023847306", "B0TK8X2B", "USB-C Cable 2m", 2, 25.98, 0, 0, 3.90, 0, 12.99, "UK", "Partial Refund", "FBT"),
    ("2025-06-23", "5791023847305", "B0TK8XK1", "Kids Smart Watch GPS", 1, 59.99, 6.00, 10.00, 9.00, 4.50, 0, "UK", "Delivered", "FBT"),
    ("2025-06-23", "5791023847303", "B0TK8X3C", "Phone Case MagSafe", 2, 37.00, 3.70, 0, 5.56, 0, 0, "UK", "Delivered", "FBT"),
]

AD_ROWS = [
    ("2025-06-23", "SP-Video-Earbuds-Summer", "Video Shopping", 48900, 956, 165.40, 1420.80, 4.1),
    ("2025-06-23", "SP-Live-FlashSale", "Live Stream", 29800, 438, 105.50, 690.00, 3.6),
    ("2025-06-23", "SP-Search-PhoneAcc", "Search Ads", 13500, 251, 45.80, 352.40, 5.0),
    ("2025-06-23", "SP-Brand-HomeKitchen", "Brand Ads", 8900, 120, 38.20, 210.00, 6.2),
]

RETURN_ROWS = [
    ("2025-06-23", "5791023847306", "B0TK8X2B", "Defective cable", 12.99, 0, "Damaged"),
]

ORDER_HEADERS = [
    "Order Date",
    "Order ID",
    "SKU",
    "Product Title",
    "Units Sold",
    "Gross Sales (USD)",
    "SKU Level Discount (USD)",
    "Order Level Discount (USD)",
    "Platform Commission (USD)",
    "Shipping Subsidy (USD)",
    "Refund Amount (USD)",
    "Ship Country",
    "Order Status",
    "Fulfillment Type",
]


def _style_header(ws, headers: list[str]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(headers[col_idx - 1]) + 2)


def create_overseas_ecommerce_excel(path: Path, meta: dict | None = None) -> Path:
    meta = meta or OVERSEAS_STORE_META
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    orders = wb.active
    orders.title = "Order Details"
    _style_header(orders, ORDER_HEADERS)
    for row in ORDER_ROWS:
        orders.append(list(row))

    ads = wb.create_sheet("Advertising")
    _style_header(
        ads,
        ["Report Date", "Campaign Name", "Ad Type", "Impressions", "Clicks", "Ad Spend (USD)", "Attributed Sales (USD)", "ROAS"],
    )
    for row in AD_ROWS:
        ads.append(list(row))

    returns = wb.create_sheet("Returns")
    _style_header(
        returns,
        ["Return Date", "Original Order ID", "SKU", "Return Reason", "Refund Value (USD)", "Restocking Fee (USD)", "Disposition"],
    )
    for row in RETURN_ROWS:
        returns.append(list(row))

    meta_sheet = wb.create_sheet("Export Meta")
    meta_sheet.append(["Field", "Value"])
    for key, value in [
        ("Platform", meta["platform"]),
        ("Store Code", meta["store_code"]),
        ("Report Date", meta["report_date"]),
    ]:
        meta_sheet.append([key, value])

    wb.save(path)
    return path


def create_settlement_excel(path: Path) -> Path:
    """B 文件：结算/调整类数据，可与订单文件做加减组合。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Settlement Adjustments"
    _style_header(
        ws,
        [
            "Transaction Date",
            "Adjustment Type",
            "Order ID",
            "Description",
            "Charge Amount (USD)",
            "Credit Amount (USD)",
        ],
    )
    rows = [
        ("2025-06-23", "Platform Fee Reversal", "", "Monthly fee rebate", 0, 15.00),
        ("2025-06-23", "Order Adjustment", "5791023847291", "Shipping compensation", 0, 3.50),
        ("2025-06-23", "Penalty", "", "Late shipment penalty", 25.00, 0),
        ("2025-06-23", "Order Adjustment", "5791023847306", "Refund processing fee", 2.50, 0),
    ]
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    return path
