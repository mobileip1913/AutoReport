"""样品单清单：模板下载、导入校验、写入 config.sample_orders。"""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models import DataSource
from app.services.ds_settings import save_ds_config
from app.services.review_import import _known_order_sku_keys, _normalize_header, _to_number

SAMPLE_COLUMNS: list[tuple[str, str, bool]] = [
    ("order_id", "Order ID", True),
    ("sku_id", "SKU ID", True),
    ("logistics", "样品单运费", False),
    ("cost", "样品单成本", False),
]

SAMPLE_TEMPLATE_HEADERS = [col[1] for col in SAMPLE_COLUMNS]

SAMPLE_FIELD_CODES = {
    "logistics": "mc_sample_logistics",
    "cost": "mc_sample_cost",
}

_HEADER_ALIASES: dict[str, str] = {}
for key, header, _ in SAMPLE_COLUMNS:
    _HEADER_ALIASES[header.lower()] = key
    _HEADER_ALIASES[header.replace(" ", "").lower()] = key
_HEADER_ALIASES["order id"] = "order_id"
_HEADER_ALIASES["skuid"] = "sku_id"
_HEADER_ALIASES["运费"] = "logistics"
_HEADER_ALIASES["样品运费"] = "logistics"
_HEADER_ALIASES["样品单运费"] = "logistics"
_HEADER_ALIASES["成本"] = "cost"
_HEADER_ALIASES["样品成本"] = "cost"
_HEADER_ALIASES["样品单成本"] = "cost"


def build_sample_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "样品单"
    ws.append([
        "说明：每行一条样品 SKU；Order ID、SKU ID 必填；样品单运费/成本填在对应列；导入后汇总至日报「样品单运费」「样品单成本」，并用于识别样品单"
    ])
    ws.append(SAMPLE_TEMPLATE_HEADERS)
    ws.append(["1234567890123456789", "9876543210", 2.5, 8])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_to_key(header: str) -> str | None:
    h = _normalize_header(header).lower()
    return _HEADER_ALIASES.get(h) or _HEADER_ALIASES.get(h.replace(" ", ""))


def parse_sample_upload(content: bytes) -> tuple[list[dict], list[str]]:
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return [], ["无法读取 Excel 文件，请使用 .xlsx 格式"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["文件为空"]

    header_row_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows[:15]):
        cells = [_normalize_header(c) for c in row]
        mapping: dict[str, int] = {}
        for j, h in enumerate(cells):
            key = _header_to_key(h)
            if key and key not in mapping:
                mapping[key] = j
        if "order_id" in mapping and "sku_id" in mapping:
            header_row_idx = i
            col_map = mapping
            break

    if header_row_idx is None:
        return [], ["缺少必填列「Order ID」和「SKU ID」（或对应中文列头）"]

    parsed: list[dict] = []
    seen: set[tuple[str, str]] = set()
    label_by_key = {k: h for k, h, _ in SAMPLE_COLUMNS}
    for line_no, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(key: str):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        oid = str(cell("order_id") or "").strip()
        sku = str(cell("sku_id") or "").strip()
        if not oid and not sku:
            continue
        if oid.startswith("1234567890") and sku == "9876543210":
            continue

        if not oid:
            errors.append(f"第 {line_no} 行：Order ID 不能为空")
            continue
        if not sku:
            errors.append(f"第 {line_no} 行：SKU ID 不能为空")
            continue

        pair = (oid, sku)
        if pair in seen:
            errors.append(f"第 {line_no} 行：Order ID + SKU ID 重复（{oid} / {sku}）")
            continue
        seen.add(pair)

        record: dict = {"order_id": oid, "sku_id": sku}
        for key, _, _required in SAMPLE_COLUMNS:
            if key in ("order_id", "sku_id"):
                continue
            raw = cell(key)
            if raw is None or str(raw).strip() == "":
                record[key] = 0.0
                continue
            num = _to_number(raw)
            text = str(raw).strip()
            if text and text not in ("0", "0.0") and num == 0.0 and not isinstance(raw, (int, float)):
                errors.append(f"第 {line_no} 行：{label_by_key[key]} 不是有效数字")
                continue
            record[key] = num
        parsed.append(record)

    if not parsed and not errors:
        errors.append("未解析到有效样品单数据")
    return parsed, errors


def sample_order_ids_from_rows(rows: list[dict]) -> list[str]:
    return sorted({str(r["order_id"]).strip() for r in rows if r.get("order_id")})


def sample_field_values(ds_config: dict) -> dict[str, float]:
    rows = ds_config.get("sample_orders") or []
    out = {code: 0.0 for code in SAMPLE_FIELD_CODES.values()}
    for row in rows:
        for key, code in SAMPLE_FIELD_CODES.items():
            out[code] += _to_number(row.get(key))
    return out


def distinct_sample_order_count(rows: list[dict]) -> int:
    return len(sample_order_ids_from_rows(rows))


def import_sample_orders(
    db: Session,
    ds: DataSource,
    content: bytes,
    *,
    strict: bool = True,
) -> dict:
    sample_rows, parse_errors = parse_sample_upload(content)
    if parse_errors and not sample_rows:
        return {"ok": False, "errors": parse_errors, "imported": 0}

    known = _known_order_sku_keys(db, ds)
    unknown = [
        f"{r['order_id']}/{r['sku_id']}"
        for r in sample_rows
        if (r["order_id"], r["sku_id"]) not in known
    ]
    errors = list(parse_errors)
    if strict and unknown:
        preview = "、".join(unknown[:5])
        suffix = f" 等 {len(unknown)} 条" if len(unknown) > 5 else ""
        errors.append(f"以下 Order ID + SKU ID 在订单主表中不存在：{preview}{suffix}")

    if errors:
        return {"ok": False, "errors": errors, "imported": 0, "unknown_count": len(unknown)}

    order_ids = sample_order_ids_from_rows(sample_rows)
    cfg = save_ds_config(db, ds, {
        "sample_orders": sample_rows,
        "sample_order_ids": order_ids,
    })
    return {
        "ok": True,
        "imported": len(sample_rows),
        "sample_order_count": len(cfg.get("sample_orders") or []),
        "sample_order_distinct": distinct_sample_order_count(sample_rows),
        "sample_logistics_total": sum(_to_number(r.get("logistics")) for r in sample_rows),
        "sample_cost_total": sum(_to_number(r.get("cost")) for r in sample_rows),
    }
