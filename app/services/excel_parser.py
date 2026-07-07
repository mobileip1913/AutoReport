from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models import DataImport, DataRow, DataSource, FieldMapping, FieldMappingPart, MappingLog

HEADER_SCAN_ROWS = 15
BATCH_SIZE = 2000


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_numeric(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def _filled_count(row: tuple) -> int:
    return sum(1 for v in (row or ()) if _normalize_header(v))


def _detect_header_row(rows: list[tuple]) -> int:
    """在前若干行中找表头行：取填充单元格最多的行（免责声明/Note/标题行通常只有 1-2 个填充格）。"""
    scan = min(HEADER_SCAN_ROWS, len(rows))
    best_idx = 0
    best_filled = -1
    for idx in range(scan):
        filled = _filled_count(rows[idx])
        if filled > best_filled:
            best_filled = filled
            best_idx = idx
    return best_idx


def _is_description_row(rows: list[tuple], header_idx: int) -> bool:
    """判断表头下一行是否为「字段说明行」（如订单表第 2 行写的是列含义而非数据）。

    依据：用更靠下的真实数据行确定哪些列是数值列，若紧跟表头的那一行在这些数值列上
    多为非数值文本，则判定为说明行。
    """
    candidate_idx = header_idx + 1
    if candidate_idx >= len(rows):
        return False
    candidate = rows[candidate_idx] or ()

    sample_indices = [header_idx + 2, header_idx + 3, header_idx + 4]
    numeric_cols: set[int] = set()
    for s_idx in sample_indices:
        if s_idx >= len(rows):
            break
        for col, val in enumerate(rows[s_idx] or ()):
            if _looks_numeric(val):
                numeric_cols.add(col)

    if not numeric_cols:
        return False

    mismatch = 0
    checked = 0
    for col in numeric_cols:
        val = candidate[col] if col < len(candidate) else None
        if val is None or str(val).strip() == "":
            continue
        checked += 1
        if not _looks_numeric(val):
            mismatch += 1

    return checked >= 2 and mismatch / checked > 0.5


def _headers_match(headers: list[str], column_header: str, aliases: list) -> bool:
    normalized_headers = {_normalize_header(h) for h in headers if h}
    for candidate in [column_header, *aliases]:
        if _normalize_header(candidate) in normalized_headers:
            return True
    return False


def parse_excel_file(
    db: Session,
    data_source: DataSource,
    file_path: Path,
    report_date: str,
    store_name: str,
    sheet_allowlist: set[str] | None = None,
) -> DataImport:
    wb = load_workbook(file_path, data_only=True, read_only=True)
    data_import = DataImport(
        data_source_id=data_source.id,
        file_name=file_path.name,
        report_date=report_date,
        store_name=store_name,
    )
    db.add(data_import)
    db.flush()

    mappings = (
        db.query(FieldMapping)
        .filter(FieldMapping.data_source_id == data_source.id)
        .all()
    )

    total_rows = 0
    pending: list[DataRow] = []

    def _flush_pending():
        nonlocal pending
        if pending:
            db.bulk_save_objects(pending)
            pending = []

    for sheet_name in wb.sheetnames:
        if sheet_allowlist is not None and sheet_name not in sheet_allowlist:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = _detect_header_row(rows)
        headers = [_normalize_header(h) for h in (rows[header_idx] or ())]
        if not any(headers):
            continue

        data_start = header_idx + 1
        if _is_description_row(rows, header_idx):
            data_start = header_idx + 2

        for mapping in mappings:
            checks: list[tuple[str, list]] = []
            for part in mapping.parts:
                if part.sheet_name == sheet_name:
                    checks.append((part.column_header, part.aliases or []))
            if not mapping.parts and mapping.sheet_name == sheet_name and mapping.column_header:
                checks.append((mapping.column_header, mapping.aliases or []))
            for col, aliases in checks:
                if not _headers_match(headers, col, aliases):
                    db.add(
                        MappingLog(
                            data_import_id=data_import.id,
                            level="warning",
                            message=f"列头未匹配: 「{mapping.logical_field.name}」期望列「{col}」",
                            context={
                                "sheet": sheet_name,
                                "expected_column": col,
                                "aliases": aliases,
                                "actual_headers": headers,
                                "logical_field": mapping.logical_field.code,
                            },
                        )
                    )

        for row_values in rows[data_start:]:
            if not any(_normalize_header(v) for v in (row_values or ())):
                continue
            row_dict = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row_values[idx] if idx < len(row_values) else None
                row_dict[header] = value
            pending.append(
                DataRow(
                    data_import_id=data_import.id,
                    sheet_name=sheet_name,
                    row_data=row_dict,
                )
            )
            total_rows += 1
            if len(pending) >= BATCH_SIZE:
                _flush_pending()

    _flush_pending()
    wb.close()

    data_import.row_count = total_rows
    db.commit()
    db.refresh(data_import)
    return data_import


def create_sample_excel(path: Path, platform: str, variant: str = "normal") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    orders = wb.active
    orders.title = "订单明细"
    if platform == "Amazon":
        headers = ["订单号", "订单日期", "商品名称", "Sales Amount", "数量", "平台费用", "Refund Amount"]
        if variant == "drift":
            headers[3] = "Revenue"
            headers[6] = "Refund"
    else:
        headers = ["订单号", "订单日期", "商品名称", "销售额", "数量", "平台费用", "退款金额"]
        if variant == "drift":
            headers[3] = "销售金额"

    orders.append(headers)
    sample_rows = [
        ["ORD-001", "2025-06-20", "蓝牙耳机", 299.0, 2, 45.0, 0],
        ["ORD-002", "2025-06-20", "手机壳", 89.0, 5, 12.0, 89.0],
        ["ORD-003", "2025-06-21", "数据线", 39.9, 10, 8.0, 0],
        ["ORD-004", "2025-06-21", "充电宝", 159.0, 3, 24.0, 0],
        ["ORD-005", "2025-06-22", "键盘", 399.0, 1, 60.0, 0],
    ]
    for row in sample_rows:
        orders.append(row)

    ads = wb.create_sheet("广告数据")
    if variant == "drift" and platform == "Amazon":
        ads.append(["日期", "Ad Spend", "点击量", "展示量"])
    else:
        ads.append(["日期", "广告花费", "点击量", "展示量"])
    ads.append(["2025-06-20", 120.5, 850, 12000])
    ads.append(["2025-06-21", 98.0, 720, 9800])
    ads.append(["2025-06-22", 145.2, 910, 13500])

    wb.save(path)
