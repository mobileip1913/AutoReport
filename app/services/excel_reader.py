"""Excel 读取：表头识别与按 Sheet 解析（供 ETL 使用，不写入 ORM）。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

HEADER_SCAN_ROWS = 15


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _filled_count(row: tuple) -> int:
    return sum(1 for v in (row or ()) if _normalize_header(v))


def detect_header_row(rows: list[tuple]) -> int:
    scan = min(HEADER_SCAN_ROWS, len(rows))
    best_idx = 0
    best_filled = -1
    for idx in range(scan):
        filled = _filled_count(rows[idx])
        if filled > best_filled:
            best_filled = filled
            best_idx = idx
    return best_idx


def _is_description_row(rows: list[tuple], header_idx: int) -> bool:
    candidate_idx = header_idx + 1
    if candidate_idx >= len(rows):
        return False

    def _looks_numeric(value: Any) -> bool:
        if value is None or value == "":
            return False
        if isinstance(value, (int, float)):
            return True
        text = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
        if not text:
            return False
        try:
            float(text)
            return True
        except ValueError:
            return False

    header = rows[header_idx]
    candidate = rows[candidate_idx]
    numeric_cols = [i for i, h in enumerate(header) if _normalize_header(h)]
    if not numeric_cols:
        return False
    sample_idx = min(header_idx + 5, len(rows) - 1)
    if sample_idx <= candidate_idx:
        return False
    sample = rows[sample_idx]
    text_hits = 0
    checked = 0
    for i in numeric_cols:
        if i >= len(sample):
            continue
        if _looks_numeric(sample[i]):
            checked += 1
            continue
        checked += 1
        if i < len(candidate) and not _looks_numeric(candidate[i]) and _normalize_header(candidate[i]):
            text_hits += 1
    return checked > 0 and text_hits / checked >= 0.5


def _is_field_description_data_row(row: tuple, headers: list[str]) -> bool:
    """TikTok 导出：表头下一行常为各列字段说明（整行是说明文字）。"""
    filled = 0
    desc_like = 0
    for i, header in enumerate(headers):
        if not header:
            continue
        val = row[i] if i < len(row) else None
        text = _normalize_header(val)
        if not text:
            continue
        filled += 1
        lower = text.lower()
        if (
            len(text) > 30
            or (text.endswith(".") and " " in text and not text[0].isdigit())
            or "unique " in lower
            or "order id" in lower and len(text) > 15
            or lower.startswith("the ")
            or " when " in lower
        ):
            desc_like += 1
    return filled >= 3 and desc_like / filled >= 0.4


def _dedupe_headers(headers: list[str]) -> list[str]:
    """重复 Excel 列头追加「 2」「 3」…，与生产 DDL `_2` 列 COMMENT 对齐。"""
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in headers:
        if not h:
            out.append("")
            continue
        count = seen.get(h, 0)
        seen[h] = count + 1
        out.append(h if count == 0 else f"{h} {count + 1}")
    return out


def read_sheet_rows(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    """返回 (headers, [{header: value}, ...])。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet 不存在: {sheet_name} in {path.name}")
    ws = wb[sheet_name]
    raw_rows: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        raw_rows.append(row)
        if len(raw_rows) >= 50000:
            break
    wb.close()

    if not raw_rows:
        return [], []

    header_idx = detect_header_row(raw_rows)

    headers = [_normalize_header(v) for v in raw_rows[header_idx]]
    while headers and not headers[-1]:
        headers.pop()
    headers = _dedupe_headers(headers)

    data_start = header_idx + 1
    while data_start < len(raw_rows) and _is_field_description_data_row(raw_rows[data_start], headers):
        data_start += 1

    records: list[dict[str, Any]] = []
    for row in raw_rows[data_start:]:
        if not any(_normalize_header(v) for v in row):
            continue
        item: dict[str, Any] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[i] if i < len(row) else None
        records.append(item)
    return headers, records
