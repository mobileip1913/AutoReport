"""日报输出：按 report_group 分组组织数值，并导出 Excel。"""

from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

from openpyxl.utils import get_column_letter

from app.config import settings

DAILY_TEMPLATE_NAME = "日报模板.xlsx"
TEMPLATE_DATA_ROW = 4
METRIC_START_COLUMN = 6  # Excel 列 F


def list_excel_templates() -> list[str]:
    """files 目录下可供店铺绑定的 Excel 导出模板（文件名含「模板」）。"""
    files_dir = Path(settings.files_dir)
    if not files_dir.is_dir():
        return [DAILY_TEMPLATE_NAME]
    names = sorted(
        p.name for p in files_dir.glob("*.xlsx")
        if p.is_file() and "模板" in p.name
    )
    if DAILY_TEMPLATE_NAME not in names:
        names.insert(0, DAILY_TEMPLATE_NAME)
    return names or [DAILY_TEMPLATE_NAME]

# 与 files/日报模板.xlsx 第 4 行列位一致；值为 ReportValue.line_label
TEMPLATE_COLUMN_LABELS: dict[str, str | None] = {
    "F": "实际支付金额",
    "G": "应支付金额",
    "H": "应收金额",
    "I": "退单金额",
    "J": "刷单金额",
    "K": "刷单佣金",
    "L": "刷单服务费",
    "M": "刷单物流费用",
    "N": "刷单成本",
    "O": "样品单运费",
    "P": "样品单成本",
    "Q": None,
    "R": "站内消耗",
    "S": None,
    "T": "下单数",
    "U": "物流费用",
    "V": "达人佣金",
    "W": "店铺佣金",
    "X": "产品成本",
    "Y": None,
    "Z": None,
    "AA": None,
    "AB": "固定费用",
    "AC": "利润",
    "AD": "框返",
    "AE": "总利润",
}

# Excel 模板预留列（系统尚未配置取数规则）
TEMPLATE_RESERVED_HINTS: dict[str, str] = {
    "Q": "实际金额",
    "S": "纯收入",
    "Y": "房租、水电",
    "Z": "人工",
    "AA": "其他",
}

EXCEL_METRIC_COLUMNS: list[tuple[str, str | None]] = [
    (col, label) for col, label in TEMPLATE_COLUMN_LABELS.items()
]

META_COLUMNS: dict[str, str] = {
    "A": "店铺名称",
    "B": "平台",
    "C": "区域",
    "D": "项目",
    "E": "日期",
}

MANUAL_EXPORT_LABELS = frozenset({"利润", "总利润", "利润(估算)", "总利润(估算)"})


def metric_col_letter(index: int) -> str:
    """按当前报表字段顺序生成 Excel 列号（从 F 起）。"""
    return get_column_letter(METRIC_START_COLUMN + index)


def column_for_report_field(mapping, mappings: list) -> str:
    """按 sort_order 从 F 列起依次编号（与日报页拖动排序、导出一致）。"""
    from app.services.mapping_utils import report_display_mappings

    ordered = report_display_mappings(mappings)
    for idx, m in enumerate(ordered):
        if m.id == mapping.id:
            return metric_col_letter(idx)
    return metric_col_letter(len(ordered))


def build_dynamic_report_rows(
    mappings: list,
    values: list | None = None,
    *,
    pending_file_codes: set[str] | frozenset[str] | None = None,
    label_fn=None,
    line_code_fn=None,
    is_manual_fn=None,
    is_formula_fn=None,
) -> list[dict]:
    """按 FieldMapping.sort_order 构建报表字段行（日报 + 报表配置共用）。"""
    from app.services.mapping_utils import (
        is_formula_line,
        is_manual_line,
        mapping_label,
        mapping_line_code,
        report_display_mappings,
    )

    label_fn = label_fn or mapping_label
    line_code_fn = line_code_fn or mapping_line_code
    is_manual_fn = is_manual_fn or is_manual_line
    is_formula_fn = is_formula_fn or is_formula_line
    pending = pending_file_codes or frozenset()

    by_mapping_id = {v.mapping_id: v for v in (values or []) if getattr(v, "mapping_id", None)}
    by_label = {v.line_label: v for v in (values or [])}
    by_code = {v.line_code: v for v in (values or []) if getattr(v, "line_code", None)}

    rows: list[dict] = []
    ordered = report_display_mappings(mappings)
    for m in ordered:
        label = label_fn(m)
        code = line_code_fn(m)
        rv = by_mapping_id.get(m.id) or by_label.get(label) or by_code.get(code)
        is_manual = is_manual_fn(m)
        computed_display = ""
        if rv and getattr(rv, "computed_raw_value", None) is not None:
            from app.services.formula import format_value
            fmt = (m.format_type or "usd")
            computed_display = format_value(rv.computed_raw_value, fmt)
        elif rv and not getattr(rv, "is_overridden", False) and rv.display_value:
            computed_display = rv.display_value

        configured = bool(m.parts or (m.sheet_name and m.column_header))
        rows.append({
            "col": column_for_report_field(m, mappings),
            "sort_order": m.sort_order or 0,
            "label": label,
            "mapping": m,
            "mapping_id": m.id,
            "line_code": code,
            "is_manual": is_manual,
            "is_formula": is_formula_fn(m),
            "is_fetch": not is_manual and not is_formula_fn(m),
            "configured": configured,
            "pending_file": code in pending,
            "format_type": m.format_type or "usd",
            "value_id": rv.id if rv else None,
            "is_overridden": bool(getattr(rv, "is_overridden", False)) if rv else False,
            "display_value": rv.display_value if rv else "",
            "raw_value": rv.raw_value if rv else None,
            "computed_display": computed_display,
            "computed_raw": getattr(rv, "computed_raw_value", None) if rv else None,
            "expression": rv.expression if rv else "",
            "editable": bool(rv),
            "is_reserved": False,
        })
    return rows


# 兼容旧调用
def build_excel_config_rows(mappings, **kwargs) -> list[dict]:
    return build_dynamic_report_rows(mappings, values=None, **kwargs)


def build_excel_report_rows(values, mappings=None, **kwargs) -> list[dict]:
    return build_dynamic_report_rows(mappings or [], values=values, **kwargs)


def build_grouped(values: list) -> list[dict]:
    """把 ReportValue 列表按 report_group 分组；无分组时按 sort_order 顺序单列展示。"""
    if not values:
        return []

    has_group = any(getattr(v, "report_group", None) for v in values)
    if not has_group:
        return [
            {
                "title": "报表指标",
                "metrics": [
                    {"label": v.line_label, "value": v.display_value, "expression": v.expression}
                    for v in sorted(values, key=lambda x: x.sort_order)
                ],
            }
        ]

    groups: list[dict] = []
    order: list[str] = []
    bucket: dict[str, list] = {}
    for v in sorted(values, key=lambda x: x.sort_order):
        title = getattr(v, "report_group", None) or "其他"
        if title not in bucket:
            bucket[title] = []
            order.append(title)
        bucket[title].append(
            {"label": v.line_label, "value": v.display_value, "expression": v.expression}
        )
    for title in order:
        groups.append({"title": title, "metrics": bucket[title]})
    return groups


def report_meta(data_source, run) -> dict:
    if data_source:
        cfg = data_source.config or {}
        meta = dict(cfg.get("meta") or {})
        meta.setdefault("平台", data_source.platform)
    else:
        meta = {}
    meta.setdefault("店铺名称", run.store_name)
    meta.setdefault("区域", "")
    meta.setdefault("项目", "")
    meta["日期"] = run.report_date
    return meta


def _template_path(data_source=None) -> Path:
    cfg = (data_source.config or {}) if data_source else {}
    chosen = (cfg.get("excel_template_file") or "").strip() or DAILY_TEMPLATE_NAME
    path = Path(settings.files_dir) / chosen
    if path.is_file():
        return path
    fallback = Path(settings.files_dir) / DAILY_TEMPLATE_NAME
    if fallback.is_file():
        return fallback
    raise FileNotFoundError(f"未找到日报模板：{chosen}")


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    """将模板表头行样式复制到数据行，保持与日报模板一致。"""
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def _export_cell_value(value) -> float | str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return value


def _write_report_value_cell(cell, rv, label: str) -> None:
    if rv is None:
        cell.value = None
        return
    if rv.raw_value is not None:
        cell.value = _export_cell_value(rv.raw_value)
        return
    if label in MANUAL_EXPORT_LABELS or rv.display_value == "":
        cell.value = None
        return
    cell.value = _export_cell_value(rv.raw_value)


def _resolve_writable_cell(ws, row: int, col: int):
    """合并单元格仅左上角可写，解析真实可写单元格。"""
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ != "MergedCell":
        return cell
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(row=min_row, column=min_col)
    return cell


def _set_metric_header(ws, col_num: int, label: str, style_src) -> None:
    """模板指标表头多在第 2 行（F2:F3 等合并），超出 AE 的追加列写第 3 行。"""
    if col_num <= 31:
        cell = _resolve_writable_cell(ws, 2, col_num)
    else:
        cell = ws.cell(row=3, column=col_num)
    cell.value = label
    if style_src.has_style:
        cell.font = copy(style_src.font)
        cell.border = copy(style_src.border)
        cell.fill = copy(style_src.fill)
        cell.number_format = copy(style_src.number_format)
        cell.alignment = copy(style_src.alignment)


def _clear_stale_append_columns(ws, first_clear_col: int, last_col: int) -> None:
    """清除旧版导出在 AF 之后写入的残留表头/数据。"""
    for col in range(first_clear_col, last_col + 1):
        for row in (2, 3, 4):
            cell = _resolve_writable_cell(ws, row, col)
            if type(cell).__name__ != "MergedCell":
                cell.value = None


def export_daily_excel(data_source, run, values: list, mappings: list | None = None) -> Path:
    """按 files/日报模板.xlsx 版式导出：保留前 3 行表头，在第 4 行填入数据。"""
    meta = report_meta(data_source, run)

    template = _template_path(data_source)
    out_dir = Path(settings.upload_dir).parent / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"美宠日报_{meta.get('日期', '')}_run{run.id}.xlsx"
    shutil.copy2(template, out_path)

    wb = load_workbook(out_path)
    ws = wb.active
    style_row = 3
    data_row = TEMPLATE_DATA_ROW
    max_col = ws.max_column

    _copy_row_style(ws, style_row, data_row, max_col)

    for col_letter, meta_key in META_COLUMNS.items():
        cell = ws[f"{col_letter}{data_row}"]
        cell.value = meta.get(meta_key, "")

    from app.services.mapping_utils import mapping_label, report_display_mappings

    if mappings is None:
        mappings = []

    by_mapping_id = {v.mapping_id: v for v in values if getattr(v, "mapping_id", None)}
    by_label = {v.line_label: v for v in values}

    def _find_value(m):
        lbl = mapping_label(m)
        return by_mapping_id.get(m.id) or by_label.get(lbl)

    ordered = report_display_mappings(mappings)
    header_row = 3
    style_src = ws.cell(row=header_row, column=31)
    first_unused_col = METRIC_START_COLUMN + len(ordered)

    if ordered:
        max_col = max(max_col, METRIC_START_COLUMN + len(ordered) - 1)
        _copy_row_style(ws, style_row, data_row, max_col)

    for idx, m in enumerate(ordered):
        col_num = METRIC_START_COLUMN + idx
        label = mapping_label(m)
        _set_metric_header(ws, col_num, label, style_src)
        data_cell = ws.cell(row=data_row, column=col_num)
        _write_report_value_cell(data_cell, _find_value(m), label)

    # 未再使用的模板列 / 旧版 AF 追加列：清掉残留表头与数据
    if ordered:
        _clear_stale_append_columns(ws, first_unused_col, max(ws.max_column, 35))

    wb.save(out_path)
    wb.close()
    return out_path
